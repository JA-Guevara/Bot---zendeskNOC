import os
import asyncio
import zipfile
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import logging
from utils.config import EXCEL_URL, DOWNLOAD_FOLDER
from pages.outlook_page import OutlookPage

# Logger para el servidor
logger_server = logging.getLogger('main')

class ExcelPage:
    def __init__(self, page):
        self.page = page

    async def download_zip_from_email(self):
        """Descarga el archivo ZIP desde el enlace obtenido de Outlook."""
        outlook_page = OutlookPage(self.page)
        download_url = await outlook_page.outlook_extraction()

        if not download_url:
            logger_server.error("⚠️ No se pudo obtener la URL de descarga.")
            return None

        try:
            # Abrir el enlace en una nueva pestaña
            new_page = await self.page.context.new_page()
            await new_page.goto(download_url)

            # Esperar a que la descarga comience
            async with new_page.expect_download() as download_info:
                # Guardar el archivo descargado
                download = await download_info.value
                if download.size == 0:
                    logger_server.error("⚠️ El archivo descargado está vacío.")
                    return None

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                zip_filename = f"data_{timestamp}.zip"  # Evitar sobrescritura
                zip_path = os.path.join(DOWNLOAD_FOLDER, zip_filename)
                await download.save_as(zip_path)

                logger_server.info(f"📁 Archivo descargado en: {zip_path}, Tamaño: {download.size} bytes")

            # Cerrar la pestaña de descarga
            await new_page.close()
            return zip_path

        except Exception as e:
            logger_server.error(f"⚠️ Error durante la descarga del archivo: {e}")
            await new_page.close()
            return None

    def extract_csv_from_zip(self, zip_path):
        """Extrae el archivo CSV del archivo ZIP descargado."""
        if zip_path is None:
            logger_server.error("⚠️ No se proporcionó una ruta de archivo ZIP válida.")
            return None

        try:
            if not os.path.exists(zip_path):
                raise FileNotFoundError(f"No se encontró el archivo ZIP: {zip_path}")

            if not zipfile.is_zipfile(zip_path):
                raise ValueError("El archivo descargado no es un ZIP válido.")

            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                csv_files = [f for f in zip_ref.namelist() if f.endswith('.csv')]
                if not csv_files:
                    raise ValueError("No se encontró ningún archivo CSV en el ZIP.")

                zip_ref.extract(csv_files[0], DOWNLOAD_FOLDER)  # Extraer CSV
                csv_path = os.path.join(DOWNLOAD_FOLDER, csv_files[0])

            logger_server.info(f"📄 Archivo CSV extraído en: {csv_path}")
            return csv_path

        except (ValueError, FileNotFoundError, zipfile.BadZipFile) as e:
            logger_server.error(f"⚠️ {e}")
            return None
        except Exception as e:
            logger_server.error(f"⚠️ Error al extraer el archivo CSV: {e}")
            return None

    def update_excel_consolidado(self, csv_path):
        """Actualiza el archivo Excel consolidado con los datos del CSV."""
        try:
            # Verificar si el CSV existe
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"No se encontró el archivo CSV: {csv_path}")

            # Leer el CSV
            df = pd.read_csv(csv_path, encoding='utf-8')  # Especificar la codificación para evitar problemas

            # Verificar si el archivo de Excel existe
            if not os.path.exists(EXCEL_URL):
                raise FileNotFoundError(f"No se encontró el archivo Excel: {EXCEL_URL}")

            # 📂 Abrir Excel desde SharePoint (EXCEL_URL)
            logger_server.info(f"📂 Abriendo archivo Excel en: {EXCEL_URL}")
            book = load_workbook(EXCEL_URL, read_only=False)  # Asegurarse de no estar en modo solo lectura
            sheet = book["consolidado"]

            # Registrar la fecha y hora de actualización
            update_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet['A1'] = "Fecha y Hora de Actualización"
            sheet['B1'] = update_time

            # Insertar datos en Excel desde el CSV
            for i, row in enumerate(df.values, start=2):  # Desde la fila 2
                for j, value in enumerate(row, start=2):  # Desde la columna B
                    sheet.cell(row=i, column=j, value=value)

            # Guardar cambios en el Excel
            book.save(EXCEL_URL)
            logger_server.info(f"✅ Archivo Excel actualizado correctamente en: {EXCEL_URL}")

        except (ValueError, FileNotFoundError, KeyError) as e:
            logger_server.error(f"⚠️ {e}")
        except Exception as e:
            logger_server.error(f"⚠️ Error al actualizar el archivo Excel: {e}")

async def main():
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()

        # Crear una instancia de ExcelPage
        excel_page = ExcelPage(page)

        # Descargar el archivo ZIP
        zip_path = await excel_page.download_zip_from_email()

        if zip_path:
            # Extraer el CSV del ZIP
            csv_path = excel_page.extract_csv_from_zip(zip_path)

            if csv_path:
                # Actualizar el archivo Excel consolidado
                excel_page.update_excel_consolidado(csv_path)

        # Cerrar el navegador
        await browser.close()

if __name__ == "__main__":
    asyncio.run(main())